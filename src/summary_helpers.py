"""Summary export helpers for get_summary_for_channel.py."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from channel_playlists import (
    playlist_order,
    resolve_playlist_selection,
)
from openpyxl import Workbook
from video_cache import (
    display_number_to_list_pos,
    is_new_display_number,
    list_pos_to_display_number,
)

SCOPE_BYPLS = "bypls"
SCOPE_ALLPLS = "allpls"
DISPLAY_NEW = "new"


@dataclass
class VideoRecord:
    channel_index: int
    display_index: int | str
    video_id: str
    title: str
    channel_name: str
    url: str
    date_text: str
    duration_bracket: str
    duration_seconds: int | None
    playlist: str
    cost_usd: float | None = None


@dataclass
class SummarySection:
    summary: str
    records: list[VideoRecord]


@dataclass
class ExportSlot:
    list_pos: int
    display_number: int


def normalize_scope(scope: str | None) -> str:
    if not scope:
        return SCOPE_BYPLS
    return scope.strip()


def scope_mode(scope: str) -> str:
    lowered = scope.lower()
    if lowered == SCOPE_ALLPLS:
        return SCOPE_ALLPLS
    if lowered == SCOPE_BYPLS:
        return SCOPE_BYPLS
    return "single"


def range_explicit(args: argparse.Namespace, default_from: int, default_to: int) -> bool:
    from_explicit = getattr(args, "from_explicit", args.from_index != default_from)
    to_explicit = getattr(args, "to_explicit", args.to_index != default_to)
    return from_explicit or to_explicit


def format_display_index(display_number: int) -> int | str:
    return DISPLAY_NEW if is_new_display_number(display_number) else display_number


def txt_index_field_width(records: list[VideoRecord]) -> int:
    """Width for the leading index column in TXT (min 3; grows with max number)."""
    numeric = [record.display_index for record in records if isinstance(record.display_index, int)]
    if not numeric:
        return 3
    return max(3, len(str(max(numeric))))


def format_txt_index_prefix(display_index: int | str, width: int) -> str:
    if display_index == DISPLAY_NEW:
        return "new"
    return str(display_index).rjust(width)


def selected_export_slots(
    length_curr: int,
    length_old: int,
    args: argparse.Namespace,
    default_from: int,
    default_to: int,
    *,
    use_new: bool,
) -> list[ExportSlot]:
    slots: list[ExportSlot] = []
    seen_positions: set[int] = set()

    if use_new:
        n_new = max(0, length_curr - length_old)
        for pos in range(n_new):
            display_number = list_pos_to_display_number(pos, length_curr, length_old)
            if is_new_display_number(display_number) and pos not in seen_positions:
                slots.append(ExportSlot(pos, display_number))
                seen_positions.add(pos)

    start = max(1, args.from_index)
    end = min(args.to_index, length_old)
    if start <= end:
        for display_number in range(start, end + 1):
            pos = display_number_to_list_pos(display_number, length_curr, length_old)
            if 0 <= pos < length_curr and pos not in seen_positions:
                slots.append(ExportSlot(pos, display_number))
                seen_positions.add(pos)

    if not use_new and not range_explicit(args, default_from, default_to):
        slots = []
        for pos in range(length_curr):
            display_number = list_pos_to_display_number(pos, length_curr, length_old)
            if display_number >= 1:
                slots.append(ExportSlot(pos, display_number))

    return slots


def format_duration_hms(total_seconds: int) -> str:
    hours, remainder = divmod(max(0, total_seconds), 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours}:{minutes:02d}:{seconds:02d}"


def records_total_duration_seconds(records: list[VideoRecord]) -> int:
    return sum(record.duration_seconds or 0 for record in records)


def selection_parenthetical(records: list[VideoRecord]) -> str:
    return (
        f"{len(records)} selected, "
        f"{format_duration_hms(records_total_duration_seconds(records))} in total"
    )


def summary_line(
    records: list[VideoRecord],
    length_curr: int,
    length_old: int,
    *,
    playlist: str | None = None,
    use_new: bool = False,
) -> str:
    suffix = f' in playlist "{playlist}"' if playlist else ""
    if not records:
        return f"Videos none selected from total {length_old} baseline / {length_curr} current{suffix}."

    selection = selection_parenthetical(records)
    new_records = [record for record in records if record.display_index == DISPLAY_NEW]
    old_records = [record for record in records if record.display_index != DISPLAY_NEW]
    old_indices = [record.channel_index for record in old_records]

    parts: list[str] = []
    if use_new and new_records and old_records:
        parts.append(f"{len(new_records)} new video(s)")
        if len(old_indices) == 1:
            parts.append(f"video {old_indices[0]} ({selection})")
        else:
            parts.append(f"videos {old_indices[0]} to {old_indices[-1]} ({selection})")
    elif use_new and new_records:
        parts.append(f"{len(new_records)} new video(s) ({selection})")
    elif len(old_indices) == 1:
        parts.append(f"video {old_indices[0]} ({selection})")
    else:
        parts.append(f"videos {old_indices[0]} to {old_indices[-1]} ({selection})")

    body = ", ".join(parts)
    return (
        f"{body} from baseline {length_old} / current {length_curr} on channel{suffix}."
    )


def summary_output_stem(
    out_dir: Path,
    *,
    now: datetime,
    args: argparse.Namespace,
    scope: str,
    plsonly: bool,
    single_playlist: dict | None,
    default_from: int,
    default_to: int,
    use_new: bool = False,
) -> Path:
    name = f"_{now.hour:02d}_{now.minute:02d}"
    if plsonly:
        name += "_pls"
    else:
        mode = scope_mode(scope)
        if mode == SCOPE_ALLPLS:
            name += "_plsall"
        elif mode == "single":
            alias = (single_playlist or {}).get("alias", "one").lstrip("#")
            name += f"_pls_{alias}"
        else:
            name += "_plsgrp"
    if use_new:
        name += "_new"
    if not plsonly and range_explicit(args, default_from, default_to):
        if getattr(args, "from_explicit", args.from_index != default_from):
            name += f"_f_{args.from_index}"
        if getattr(args, "to_explicit", args.to_index != default_to):
            name += f"_t_{args.to_index}"
        if getattr(args, "range_end_source", None) == "next":
            next_count = getattr(args, "next_count", None)
            if next_count is not None:
                name += f"_n_{next_count}"
    return out_dir / name


def build_record(
    entry: dict,
    *,
    channel_index: int,
    display_index: int | str,
    channel_name: str,
    playlist_map: dict[str, str],
    today: date,
    should_include,
    resolve_date,
    duration_text_to_seconds,
    parse_duration_text,
) -> VideoRecord | None:
    duration_seconds = duration_text_to_seconds(entry.get("duration_text"))
    date_text = resolve_date(entry["title"], entry.get("relative_published"), today)
    if not should_include(entry["title"], date_text, duration_seconds):
        return None
    return VideoRecord(
        channel_index=channel_index,
        display_index=display_index,
        video_id=entry["id"],
        title=entry["title"].strip(),
        channel_name=channel_name,
        url=entry["url"],
        date_text=date_text,
        duration_bracket=parse_duration_text(entry.get("duration_text")),
        duration_seconds=duration_seconds,
        playlist=playlist_map.get(entry["id"], ""),
    )


def build_records_for_slots(
    browse_videos: list[dict],
    channel_name: str,
    slots: list[ExportSlot],
    playlist_map: dict[str, str],
    today: date,
    *,
    numbering: str,
    should_include,
    resolve_date,
    duration_text_to_seconds,
    parse_duration_text,
) -> tuple[list[VideoRecord], int, int]:
    total_channel = len(browse_videos)
    records: list[VideoRecord] = []
    excluded = 0
    playlist_counter = 0
    for slot in slots:
        if slot.list_pos < 0 or slot.list_pos >= total_channel:
            continue
        entry = browse_videos[slot.list_pos]
        display_index = format_display_index(slot.display_number)
        if numbering == "playlist" and not is_new_display_number(slot.display_number):
            playlist_counter += 1
            display_index = playlist_counter
        elif numbering == "playlist" and is_new_display_number(slot.display_number):
            display_index = DISPLAY_NEW
        record = build_record(
            entry,
            channel_index=slot.display_number,
            display_index=display_index,
            channel_name=channel_name,
            playlist_map=playlist_map,
            today=today,
            should_include=should_include,
            resolve_date=resolve_date,
            duration_text_to_seconds=duration_text_to_seconds,
            parse_duration_text=parse_duration_text,
        )
        if record is None:
            excluded += 1
            continue
        records.append(record)
    return records, total_channel, excluded


def filter_slots_for_playlist(
    slots: list[ExportSlot],
    browse_videos: list[dict],
    playlist_map: dict[str, str],
    playlist_title: str,
) -> list[ExportSlot]:
    return [
        slot
        for slot in slots
        if slot.list_pos < len(browse_videos)
        and playlist_map.get(browse_videos[slot.list_pos]["id"], "") == playlist_title
    ]


def build_summary_sections(
    browse_videos: list[dict],
    channel_name: str,
    export_slots: list[ExportSlot],
    playlist_map: dict[str, str],
    playlists_cache: dict,
    scope: str,
    today: date,
    args: argparse.Namespace,
    default_from: int,
    default_to: int,
    length_old: int,
    *,
    use_new: bool,
    should_include,
    resolve_date,
    duration_text_to_seconds,
    parse_duration_text,
) -> tuple[list[SummarySection], int, int]:
    mode = scope_mode(scope)
    total = len(browse_videos)
    excluded_total = 0

    if mode == "single":
        playlist = resolve_playlist_selection(scope, playlists_cache)
        pl_slots = filter_slots_for_playlist(
            export_slots, browse_videos, playlist_map, playlist["title"]
        )
        records, _, excluded = build_records_for_slots(
            browse_videos,
            channel_name,
            pl_slots,
            playlist_map,
            today,
            numbering="channel",
            should_include=should_include,
            resolve_date=resolve_date,
            duration_text_to_seconds=duration_text_to_seconds,
            parse_duration_text=parse_duration_text,
        )
        excluded_total += excluded
        return (
            [
                SummarySection(
                    summary_line(
                        records,
                        total,
                        length_old,
                        playlist=playlist["title"],
                        use_new=use_new,
                    ),
                    records,
                )
            ],
            total,
            excluded_total,
        )

    if mode == SCOPE_ALLPLS:
        records, _, excluded = build_records_for_slots(
            browse_videos,
            channel_name,
            export_slots,
            playlist_map,
            today,
            numbering="channel",
            should_include=should_include,
            resolve_date=resolve_date,
            duration_text_to_seconds=duration_text_to_seconds,
            parse_duration_text=parse_duration_text,
        )
        excluded_total += excluded
        return (
            [
                SummarySection(
                    summary_line(records, total, length_old, use_new=use_new),
                    records,
                )
            ],
            total,
            excluded_total,
        )

    use_playlist_numbering = not range_explicit(args, default_from, default_to) and not use_new
    sections: list[SummarySection] = []
    known_titles = {pl["title"] for pl in playlist_order(playlists_cache)}

    for playlist in playlist_order(playlists_cache):
        pl_slots = filter_slots_for_playlist(
            export_slots, browse_videos, playlist_map, playlist["title"]
        )
        if not pl_slots:
            continue
        records, _, excluded = build_records_for_slots(
            browse_videos,
            channel_name,
            pl_slots,
            playlist_map,
            today,
            numbering="playlist" if use_playlist_numbering else "channel",
            should_include=should_include,
            resolve_date=resolve_date,
            duration_text_to_seconds=duration_text_to_seconds,
            parse_duration_text=parse_duration_text,
        )
        excluded_total += excluded
        sections.append(
            SummarySection(
                summary_line(
                    records,
                    total,
                    length_old,
                    playlist=playlist["title"],
                    use_new=use_new,
                ),
                records,
            )
        )

    other_slots = [
        slot
        for slot in export_slots
        if slot.list_pos < total
        and playlist_map.get(browse_videos[slot.list_pos]["id"], "") not in known_titles
    ]
    if other_slots:
        records, _, excluded = build_records_for_slots(
            browse_videos,
            channel_name,
            other_slots,
            playlist_map,
            today,
            numbering="playlist" if use_playlist_numbering else "channel",
            should_include=should_include,
            resolve_date=resolve_date,
            duration_text_to_seconds=duration_text_to_seconds,
            parse_duration_text=parse_duration_text,
        )
        excluded_total += excluded
        sections.append(
            SummarySection(
                summary_line(
                    records,
                    total,
                    length_old,
                    playlist="(no playlist)",
                    use_new=use_new,
                ),
                records,
            )
        )

    return sections, total, excluded_total


def write_plsonly_txt(path: Path, header: str, rows: list[tuple[str, str, str]]) -> None:
    lines = [header, *(f"{alias}\t{channel}\t{title}" for alias, channel, title in rows)]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_plsonly_xlsx(path: Path, header: str, rows: list[tuple[str, str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Playlists"
    ws.cell(row=1, column=1, value=header)
    row = 2
    for alias, channel, title in rows:
        ws.cell(row=row, column=1, value=alias)
        ws.cell(row=row, column=2, value=channel)
        ws.cell(row=row, column=3, value=title)
        row += 1
    wb.save(path)


def format_txt_cost(cost_usd: float | None, width: int) -> str:
    text = "$?" if cost_usd is None else f"${cost_usd:.2f}"
    return text.rjust(width)


def txt_cost_field_width(records: list[VideoRecord]) -> int:
    widths = [
        len("$?" if record.cost_usd is None else f"${record.cost_usd:.2f}")
        for record in records
    ]
    return max(widths, default=2)


def write_summary_txt(
    path: Path,
    sections: list[SummarySection],
    *,
    format_txt_line,
) -> None:
    all_records = [record for section in sections for record in section.records]
    index_width = txt_index_field_width(all_records)
    cost_width = txt_cost_field_width(all_records)
    lines: list[str] = []
    for index, section in enumerate(sections):
        if index:
            lines.append("")
        lines.append(section.summary)
        for record in section.records:
            prefix = format_txt_index_prefix(record.display_index, index_width)
            cost = format_txt_cost(record.cost_usd, cost_width)
            lines.append(f"{prefix} {cost} {format_txt_line(record)}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_summary_xlsx(
    path: Path,
    sections: list[SummarySection],
    *,
    parse_excel_date,
    duration_bracket_to_seconds,
    build_display_title,
    split_display_title,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Videos"
    row = 1
    for section_index, section in enumerate(sections):
        if section_index:
            row += 1
        ws.cell(row=row, column=1, value=section.summary)
        row += 1
        for record in section.records:
            parsed_date = parse_excel_date(record.date_text)
            ws.cell(row=row, column=1, value=record.display_index)
            if record.cost_usd is not None:
                cost_cell = ws.cell(row=row, column=2, value=record.cost_usd)
                cost_cell.number_format = '"$"0.00'
            ws.cell(row=row, column=3, value=record.channel_name)
            ws.cell(row=row, column=4, value=record.playlist)
            ws.cell(row=row, column=5, value=record.url)
            date_cell = ws.cell(row=row, column=6, value=parsed_date)
            if parsed_date is not None:
                date_cell.number_format = (
                    "yyyy-mm-dd" if len(record.date_text) == 10 else "yyyy-mm"
                )
            duration_seconds = record.duration_seconds or duration_bracket_to_seconds(
                record.duration_bracket
            )
            if duration_seconds is not None:
                duration_cell = ws.cell(row=row, column=7, value=duration_seconds / 86400)
                duration_cell.number_format = "[h]:mm:ss"
            display_title = build_display_title(record.title, record.channel_name)
            title_only, _ = split_display_title(display_title, record.channel_name)
            ws.cell(row=row, column=8, value=title_only)
            row += 1
    wb.save(path)
