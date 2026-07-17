#!/usr/bin/env python3
"""Export YouTube channel video summaries to TXT and XLSX (pipeline step 1).

Examples:
  python src/export_channel.py @Ekaterina_Schulmann
  python src/export_channel.py UCL1rJ0ROIw9V1qFeIN0ZTZQ --new
  python src/export_channel.py @VladilenMinin Vladilen_Minin --from 1 --to 50
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

_SRC_DIR = Path(__file__).resolve().parent
if str(_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(_SRC_DIR))

from openpyxl import Workbook

from playlist_mapping import build_video_playlist_map
from project_paths import WORKSPACE_ROOT, cache_dir, output_dir

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

API_KEY = "AIzaSyAO_FJ2SlqU8Q4STEHLGCilw_Y9_11qcW8"
WAR_START = date(2022, 2, 24)
STREAM_CUTOFF = date(2020, 1, 1)
DEFAULT_FROM = 1
DEFAULT_TO = 10000
CHANNEL_ID_RE = re.compile(r"^UC[\w-]{22}$")
INVALID_PATH_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
CACHE_HEAD_CHECK = 3
PAGE_SIZE = 30  # для оценки номера страницы по индексу видео (--to); API может вернуть 28–30
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
BROWSE_ID_RE = re.compile(r'"browseId"\s*:\s*"(UC[\w-]{22})"')
CHANNEL_URL_ID_RE = re.compile(r"channel/(UC[\w-]{22})")
CANONICAL_HANDLE_RE = re.compile(r'"canonicalBaseUrl"\s*:\s*"(/@[^"]+)"')

DAY_PATTERNS = [
    re.compile(r"(?:День|ДЕНЬ|день)\s*(\d+)", re.IGNORECASE),
    re.compile(r"(\d+)\s*[-–—]?\s*(?:й|Й)\s*(?:день|ДЕНЬ)", re.IGNORECASE),
    re.compile(r"(?:^|[\s|])ДЕНЬ\s+(\d+)", re.IGNORECASE),
    re.compile(r"(?:^|[\s.])(\d{3,5})\.\s", re.IGNORECASE),
]
STREAM_HINTS = re.compile(r"(?:стрим|stream|эфир|live\b)", re.IGNORECASE)
RELATIVE_RE = re.compile(
    r"(?P<num>\d+)\s*(?P<unit>second|minute|hour|day|week|month|year)s?\s+ago",
    re.IGNORECASE,
)


@dataclass
class VideoRecord:
    channel_index: int
    video_id: str
    title: str
    channel_name: str
    url: str
    date_text: str
    duration_bracket: str
    duration_seconds: int | None
    playlist: str


@dataclass
class FetchResult:
    videos: list[dict]
    channel_name: str
    new_video_ids: set[str]
    cache_used: bool



def yt_dlp_run(args: argparse.Namespace, *extra: str) -> subprocess.CompletedProcess[str]:
    cmd = [args.yt_dlp, *extra]
    if args.cookies_from_browser:
        cmd[1:1] = ["--cookies-from-browser", args.cookies_from_browser]
    return subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )


def folder_name_from_handle(handle: str) -> str:
    name = handle.strip()
    if name.startswith("@"):
        name = name[1:]
    name = name.replace(" ", "_")
    name = INVALID_PATH_CHARS.sub("", name)
    name = name.strip(" .")
    return name or "unnamed_channel"


def normalize_handle(value: str) -> str:
    value = value.strip()
    if value.startswith("https://www.youtube.com/"):
        value = value.rstrip("/").split("/")[-1]
    if not value.startswith("@"):
        value = "@" + value
    return value


def fetch_youtube_page(url: str) -> str:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept-Language": "en-US,en;q=0.9",
        },
    )
    with urllib.request.urlopen(req, timeout=45) as resp:
        return resp.read().decode("utf-8", "replace")


def extract_browse_id(html: str) -> str | None:
    match = BROWSE_ID_RE.search(html)
    if match:
        return match.group(1)
    match = CHANNEL_URL_ID_RE.search(html)
    return match.group(1) if match else None


def extract_canonical_handle(html: str) -> str | None:
    match = CANONICAL_HANDLE_RE.search(html)
    if not match:
        return None
    return normalize_handle(match.group(1).lstrip("/"))


def resolve_channel_id(channel_arg: str, args: argparse.Namespace) -> tuple[str, str | None]:
    raw = channel_arg.strip()
    handle: str | None = None

    if raw.startswith("@"):
        handle = normalize_handle(raw)
        channel_url = f"https://www.youtube.com/{handle}"
        html = fetch_youtube_page(channel_url)
        channel_id = extract_browse_id(html)
        if not channel_id:
            raise SystemExit(f"Could not resolve channel id for {handle}.")
        return channel_id, handle

    if CHANNEL_ID_RE.fullmatch(raw):
        return raw, None

    if raw.startswith("http"):
        html = fetch_youtube_page(raw)
        channel_id = extract_browse_id(html)
        if not channel_id:
            raise SystemExit(f"Could not resolve channel id from URL: {raw}")
        handle = extract_canonical_handle(html)
        return channel_id, handle

    raise SystemExit(
        "First argument must be a channel id (UC…), @handle, or channel URL."
    )


def resolve_channel_handle(channel_id: str, args: argparse.Namespace) -> str | None:
    html = fetch_youtube_page(f"https://www.youtube.com/channel/{channel_id}")
    return extract_canonical_handle(html)


def resolve_output_folder(
    channel_id: str,
    handle: str | None,
    explicit: str | None,
    channel_name: str,
    args: argparse.Namespace,
) -> str:
    if explicit:
        return explicit
    if handle:
        return folder_name_from_handle(handle)
    resolved = resolve_channel_handle(channel_id, args)
    if resolved:
        return folder_name_from_handle(resolved)
    slug = re.sub(r"\W+", "_", channel_name.strip().lower()).strip("_")
    return slug[:60] or channel_id


def post_innertube(endpoint: str, payload: dict, retries: int = 5) -> dict:
    data = json.dumps(payload).encode("utf-8")
    url = f"https://www.youtube.com/youtubei/v1/{endpoint}?key={API_KEY}"
    req = urllib.request.Request(
        url,
        data=data,
        headers={
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept-Language": "en-US,en;q=0.9",
        },
        method="POST",
    )
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(req, timeout=45) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            if exc.code in {429, 500, 503} and attempt + 1 < retries:
                time.sleep(2**attempt)
                continue
            raise
    raise RuntimeError(f"Failed POST {endpoint}")


def client_context() -> dict:
    return {
        "clientName": "WEB",
        "clientVersion": "2.20240410.01.00",
        "hl": "en",
        "gl": "US",
    }


def find_continuations(obj, tokens: list[str]) -> None:
    if isinstance(obj, dict):
        cmd = obj.get("continuationCommand")
        if isinstance(cmd, dict) and cmd.get("token"):
            tokens.append(cmd["token"])
        for value in obj.values():
            find_continuations(value, tokens)
    elif isinstance(obj, list):
        for item in obj:
            find_continuations(item, tokens)


def extract_channel_name(body: dict, fallback: str = "YouTube Channel") -> str:
    metadata = body.get("metadata", {}).get("channelMetadataRenderer", {})
    title = metadata.get("title")
    if isinstance(title, str) and title.strip():
        return title.strip()
    return fallback


def extract_lockups(body: dict) -> list[dict]:
    lockups: list[dict] = []

    def walk(obj) -> None:
        if isinstance(obj, dict):
            if "lockupViewModel" in obj:
                lockups.append(obj["lockupViewModel"])
            for value in obj.values():
                walk(value)
        elif isinstance(obj, list):
            for item in obj:
                walk(item)

    walk(body)
    return lockups


def parse_duration_text(text: str | None) -> str:
    if not text:
        return "[?:??]"
    text = text.strip()
    if re.fullmatch(r"\d+:\d{2}:\d{2}", text):
        hours, minutes, seconds = text.split(":")
        return f"[{int(hours)}:{int(minutes):02d}:{int(seconds):02d}]"
    if re.fullmatch(r"\d+:\d{2}", text):
        minutes, seconds = text.split(":")
        return f"[{int(minutes)}:{int(seconds):02d}]"
    return f"[{text}]"


def duration_text_to_seconds(text: str | None) -> int | None:
    if not text:
        return None
    text = text.strip()
    if re.fullmatch(r"\d+:\d{2}:\d{2}", text):
        hours, minutes, seconds = map(int, text.split(":"))
        return hours * 3600 + minutes * 60 + seconds
    if re.fullmatch(r"\d+:\d{2}", text):
        minutes, seconds = map(int, text.split(":"))
        return minutes * 60 + seconds
    return None


def duration_bracket_to_seconds(duration: str) -> int | None:
    match = re.fullmatch(r"\[(\d+):(\d{2})(?::(\d{2}))?\]", duration)
    if not match:
        return None
    if match.group(3) is not None:
        return int(match.group(1)) * 3600 + int(match.group(2)) * 60 + int(match.group(3))
    return int(match.group(1)) * 60 + int(match.group(2))


def war_day_from_title(title: str) -> date | None:
    for pattern in DAY_PATTERNS:
        match = pattern.search(title)
        if not match:
            continue
        day_num = int(match.group(1))
        if 1 <= day_num <= 5000:
            return WAR_START + timedelta(days=day_num - 1)
    return None


def parse_relative_date(text: str | None, today: date) -> str:
    if not text:
        return "????-??-??"
    lowered = text.strip().lower()
    if lowered in {"streamed live", "premiere"}:
        return "????-??-??"
    match = RELATIVE_RE.search(lowered)
    if not match:
        return "????-??-??"
    num = int(match.group("num"))
    unit = match.group("unit").lower()
    if unit in {"month", "year"}:
        if unit == "month":
            month_index = today.year * 12 + today.month - num
        else:
            month_index = today.year * 12 + today.month - num * 12
        year, month = divmod(month_index - 1, 12)
        month += 1
        return f"{year:04d}-{month:02d}"
    delta_map = {
        "second": timedelta(seconds=num),
        "minute": timedelta(minutes=num),
        "hour": timedelta(hours=num),
        "day": timedelta(days=num),
        "week": timedelta(weeks=num),
    }
    delta = delta_map.get(unit)
    if not delta:
        return "????-??-??"
    return (today - delta).isoformat()


def is_before_2020(resolved_date: str) -> bool:
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", resolved_date):
        return date.fromisoformat(resolved_date) < STREAM_CUTOFF
    if re.fullmatch(r"\d{4}-\d{2}", resolved_date):
        return date(int(resolved_date[:4]), int(resolved_date[5:7]), 1) < STREAM_CUTOFF
    return False


def is_stream(title: str, duration_seconds: int | None) -> bool:
    if STREAM_HINTS.search(title):
        return True
    if duration_seconds and duration_seconds >= 3600 and re.search(
        r"в гостях|Плющев и Наки|Плющева и Наки|Плющева с Наки",
        title,
        re.IGNORECASE,
    ):
        return True
    if duration_seconds and duration_seconds >= 7200 and war_day_from_title(title):
        return True
    return False


def resolve_date(title: str, relative_published: str | None, today: date) -> str:
    war_day = war_day_from_title(title)
    if war_day:
        return war_day.isoformat()
    return parse_relative_date(relative_published, today)


def should_include(title: str, resolved_date: str, duration_seconds: int | None) -> bool:
    return not (is_before_2020(resolved_date) and is_stream(title, duration_seconds))


def parse_lockup(lockup: dict) -> dict | None:
    video_id = lockup.get("contentId")
    if not video_id:
        return None
    meta = lockup.get("metadata", {}).get("lockupMetadataViewModel", {})
    title = (meta.get("title", {}) or {}).get("content", "").strip()
    if not title:
        return None

    relative_published = None
    rows = (
        meta.get("metadata", {})
        .get("contentMetadataViewModel", {})
        .get("metadataRows", [])
        or []
    )
    for row in rows:
        for part in row.get("metadataParts", []) or []:
            text_obj = part.get("text", {}) or {}
            text = text_obj.get("content") or text_obj.get("simpleText")
            if text and "view" not in text.lower() and "watching" not in text.lower():
                relative_published = text
                break
        if relative_published:
            break

    duration_text = None
    overlays = (
        lockup.get("contentImage", {})
        .get("thumbnailViewModel", {})
        .get("overlays", [])
        or []
    )
    for overlay in overlays:
        badges = overlay.get("thumbnailBottomOverlayViewModel", {}).get("badges", []) or []
        for badge in badges:
            text = badge.get("thumbnailBadgeViewModel", {}).get("text")
            if text and re.search(r"\d", text):
                duration_text = text
                break
        if duration_text:
            break

    return {
        "id": video_id,
        "title": title,
        "duration_text": duration_text,
        "relative_published": relative_published,
        "url": f"https://www.youtube.com/watch?v={video_id}",
    }


def cache_file_path(channel_id: str, workspace: Path) -> Path:
    return cache_dir(workspace) / f"{channel_id}.json"


def read_cache(path: Path) -> dict | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None


def write_cache(
    path: Path,
    *,
    channel_id: str,
    channel_name: str,
    channel_handle: str | None,
    videos: list[dict],
    continuation_token: str | None = None,
    pages_fetched: int | None = None,
    playlist_map: dict[str, str] | None = None,
    keep_playlist_map: bool = False,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = read_cache(path) if keep_playlist_map else None
    payload = {
        "channel_id": channel_id,
        "channel_name": channel_name,
        "channel_handle": channel_handle,
        "count": len(videos),
        "last_fetched_at": datetime.now().isoformat(timespec="seconds"),
        "continuation_token": continuation_token,
        "pages_fetched": pages_fetched,
        "videos": videos,
    }
    if playlist_map is not None:
        payload["playlist_map"] = playlist_map
        payload["playlist_map_at"] = datetime.now().isoformat(timespec="seconds")
    elif existing and existing.get("playlist_map"):
        payload["playlist_map"] = existing["playlist_map"]
        payload["playlist_map_at"] = existing.get("playlist_map_at")
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def fetch_channel_name(channel_id: str) -> str:
    payload = {
        "context": {"client": client_context()},
        "browseId": channel_id,
        "params": "EgZ2aWRlb3PyBgQKAjoA",
    }
    body = post_innertube("browse", payload)
    return extract_channel_name(body)


def fetch_browse_page(channel_id: str, continuation: str | None = None) -> dict:
    if continuation:
        payload = {"context": {"client": client_context()}, "continuation": continuation}
    else:
        payload = {
            "context": {"client": client_context()},
            "browseId": channel_id,
            "params": "EgZ2aWRlb3PyBgQKAjoA",
        }
    return post_innertube("browse", payload)


def first_continuation(body: dict) -> str | None:
    tokens: list[str] = []
    find_continuations(body, tokens)
    return tokens[0] if tokens else None


def ingest_lockups_from_response(
    response: dict,
    videos: list[dict],
    seen_ids: set[str],
) -> tuple[int, int]:
    lockups = extract_lockups(response)
    added = 0
    for lockup in lockups:
        parsed = parse_lockup(lockup)
        if parsed and parsed["id"] not in seen_ids:
            seen_ids.add(parsed["id"])
            videos.append(parsed)
            added += 1
    return added, len(lockups)


def log_page_ingest(
    page: int,
    total: int,
    added: int,
    items: int,
    *,
    is_last: bool = False,
) -> None:
    """Log one browse page. `added` = new videos on this page; no padding to PAGE_SIZE."""
    if is_last:
        suffix = ", last page"
    elif items and added == 0 and total == 0:
        suffix = f", 0 parsed from {items} lockups"
    elif added == 0:
        suffix = ", all already in cache"
    elif added < items:
        suffix = f", {items - added} skipped"
    else:
        suffix = ""
    print(f"Page {page}: total {total} (+{added}{suffix})", flush=True)


def browse_api_header() -> None:
    print("Fetching channel videos via YouTube browse API...", flush=True)


def fetch_all_browse_videos(channel_id: str) -> tuple[list[dict], str, str | None, int]:
    browse_api_header()
    body = fetch_browse_page(channel_id)
    channel_name = extract_channel_name(body)
    videos: list[dict] = []
    seen_ids: set[str] = set()
    added, items = ingest_lockups_from_response(body, videos, seen_ids)
    log_page_ingest(1, len(videos), added, items)

    token = first_continuation(body)
    page = 1

    while token:
        page += 1
        response = fetch_browse_page(channel_id, token)
        added, items = ingest_lockups_from_response(response, videos, seen_ids)
        token = first_continuation(response)
        if items == 0:
            log_page_ingest(page, len(videos), added, items, is_last=True)
            break
        log_page_ingest(page, len(videos), added, items, is_last=not token)
        time.sleep(0.4)

    print(
        f"Channel browse complete: {len(videos)} videos in {page} page(s).",
        flush=True,
    )
    return videos, channel_name, None, page


def pages_for_index(index: int) -> int:
    return max(1, (index - 1) // PAGE_SIZE + 1)


def fetch_first_browse_page(
    channel_id: str,
    *,
    known_ids: set[str] | None = None,
    announce: bool = True,
) -> tuple[list[dict], str, dict]:
    if announce:
        browse_api_header()
    body = fetch_browse_page(channel_id)
    channel_name = extract_channel_name(body)
    lockups = extract_lockups(body)
    videos: list[dict] = []
    added = 0
    for lockup in lockups:
        parsed = parse_lockup(lockup)
        if not parsed:
            continue
        videos.append(parsed)
        if known_ids is None or parsed["id"] not in known_ids:
            added += 1
    log_page_ingest(1, len(videos), added, len(lockups))
    return videos, channel_name, body


def fetch_first_page_videos(channel_id: str) -> tuple[list[dict], str]:
    videos, channel_name, _ = fetch_first_browse_page(channel_id)
    return videos, channel_name


def fetch_browse_up_to_count(
    channel_id: str,
    min_count: int | None,
    existing: list[dict],
    first_body: dict | None = None,
    *,
    resume_token: str | None = None,
    resume_after_page: int = 0,
) -> tuple[list[dict], str, str | None, int]:
    """Fetch browse pages until min_count videos are cached, or all pages if min_count is None.

    Returns (videos, channel_name, continuation_token_for_next_page, pages_fetched).
    When resume_token is set, skips network requests for pages already in cache.
    """
    videos = list(existing)
    seen_ids = {v["id"] for v in videos}

    if min_count is not None and len(videos) >= min_count:
        if first_body is not None and resume_after_page == 0 and resume_token is None:
            return (
                videos,
                extract_channel_name(first_body),
                first_continuation(first_body) if len(videos) >= PAGE_SIZE else None,
                max(1, pages_fetched or 1),
            )
        channel_name = (
            extract_channel_name(first_body) if first_body else fetch_channel_name(channel_id)
        )
        return videos, channel_name, resume_token, resume_after_page

    page = resume_after_page
    pages_fetched = resume_after_page
    token: str | None = None
    channel_name = ""

    if resume_token:
        channel_name = (
            extract_channel_name(first_body) if first_body else fetch_channel_name(channel_id)
        )
        print(
            f"Resuming from page {page + 1} ({len(videos)} videos already in cache, "
            f"no re-fetch of page{'s' if page > 1 else ''} "
            f"{'1-' + str(page) if page > 1 else '1'})...",
            flush=True,
        )
        token = resume_token
    else:
        if first_body is not None:
            body = first_body
        else:
            browse_api_header()
            body = fetch_browse_page(channel_id)
        channel_name = extract_channel_name(body)
        page = 1
        pages_fetched = 1

        if len(videos) < PAGE_SIZE:
            added, items = ingest_lockups_from_response(body, videos, seen_ids)
            if page == 1 and not resume_token:
                log_page_ingest(1, len(videos), added, items)
            if added == 0 and items > 0 and len(videos) == 0:
                raise SystemExit(
                    f"Could not parse any videos from page 1 ({items} lockups on page)."
                )

        if min_count is not None and len(videos) >= min_count:
            return videos, channel_name, first_continuation(body), pages_fetched

        token = first_continuation(body)
        if token and len(videos) >= PAGE_SIZE and min_count is not None:
            print(
                f"Walking from page 2 ({len(videos)} videos in cache, "
                f"no saved continuation token)...",
                flush=True,
            )

    while token and (min_count is None or len(videos) < min_count):
        page += 1
        response = fetch_browse_page(channel_id, token)
        added, items = ingest_lockups_from_response(response, videos, seen_ids)
        pages_fetched = page
        token = first_continuation(response)
        if items == 0:
            log_page_ingest(page, len(videos), added, items, is_last=True)
            print("Empty browse page; channel list complete.", flush=True)
            token = None
            break
        log_page_ingest(page, len(videos), added, items, is_last=not token)
        if min_count is not None and len(videos) >= min_count:
            break
        if token:
            time.sleep(0.4)

    if min_count is None and pages_fetched:
        print(
            f"Channel browse complete: {len(videos)} videos in {pages_fetched} page(s).",
            flush=True,
        )

    next_token = token if token else None
    return videos, channel_name, next_token, pages_fetched


def new_videos_at_head(fresh: list[dict], cached: list[dict]) -> list[dict]:
    cached_ids = {v["id"] for v in cached}
    new: list[dict] = []
    for video in fresh:
        if video["id"] not in cached_ids:
            new.append(video)
        else:
            break
    return new


def cache_head_matches(fresh: list[dict], cached: list[dict], n: int = CACHE_HEAD_CHECK) -> bool:
    limit = min(n, len(fresh), len(cached))
    if limit == 0:
        return False
    for i in range(limit):
        if fresh[i]["id"] != cached[i]["id"]:
            return False
        if fresh[i].get("title") != cached[i].get("title"):
            return False
    return True


def merge_new_with_cache(new_videos: list[dict], cached_videos: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {v["id"]: v for v in cached_videos}
    for video in reversed(new_videos):
        merged[video["id"]] = video
    ordered: list[dict] = []
    seen: set[str] = set()
    for video in new_videos:
        if video["id"] not in seen:
            ordered.append(video)
            seen.add(video["id"])
    for video in cached_videos:
        if video["id"] not in seen:
            ordered.append(video)
            seen.add(video["id"])
    return ordered


def needs_head_check(
    args: argparse.Namespace,
    cached_len: int,
    required: int | None,
    cached_total: int,
    continuation_token: str | None,
) -> bool:
    """Whether page 1 must be fetched from YouTube before using/extending cache."""
    if args.new:
        return True
    if required is not None and cached_len >= required:
        return False
    if required is None:
        if continuation_token is None and cached_total > 0 and cached_len >= cached_total:
            return False
        return True
    if range_explicit(args) and args.from_index > CACHE_HEAD_CHECK:
        return False
    return True


def cache_status_line(cached_len: int, cache: dict, required: int | None) -> str:
    pages = int(cache.get("pages_fetched") or 0)
    pages_part = f", pages 1-{pages} in cache" if pages else ""
    need_part = f", need {required}" if required is not None else ", need full list"
    return f"{cached_len} videos{pages_part}{need_part}"


def incremental_smart_load(
    channel_id: str,
    args: argparse.Namespace,
    cache: dict,
    handle: str | None,
    cache_path: Path,
) -> FetchResult:
    """Page-aware smart refresh (default mode and --new with optional range)."""
    cached_videos: list[dict] = list(cache["videos"])
    channel_name = cache.get("channel_name") or ""
    cached_total = cache.get("count", len(cached_videos))

    if args.new and not range_explicit(args):
        # --new without range always checks channel head
        print("Smart refresh: checking channel head (page 1)...", flush=True)
        browse_api_header()
        cached_ids = {v["id"] for v in cached_videos}
        first_page, channel_name, first_body = fetch_first_browse_page(
            channel_id, known_ids=cached_ids, announce=False
        )
        if cached_videos and not cache_head_matches(first_page, cached_videos):
            print("Cache head mismatch; performing full refresh.", flush=True)
            videos, channel_name, _, pages_fetched = fetch_all_browse_videos(channel_id)
            write_cache(
                cache_path,
                channel_id=channel_id,
                channel_name=channel_name,
                channel_handle=handle or cache.get("channel_handle"),
                videos=videos,
                continuation_token=None,
                pages_fetched=pages_fetched,
            )
            return FetchResult(videos, channel_name, set(), False)
        new_head = new_videos_at_head(first_page, cached_videos) if cached_videos else []
        new_ids = {v["id"] for v in new_head}
        if new_head:
            cached_videos = merge_new_with_cache(new_head, cached_videos)
            write_cache(
                cache_path,
                channel_id=channel_id,
                channel_name=channel_name,
                channel_handle=handle or cache.get("channel_handle"),
                videos=cached_videos,
                continuation_token=None,
                pages_fetched=1,
                keep_playlist_map=True,
            )
        else:
            print("No new videos found.", flush=True)
        return FetchResult(cached_videos, channel_name, new_ids, True)

    if range_explicit(args):
        required = args.to_index
    elif cache.get("continuation_token") is None and cached_total > 0 and len(cached_videos) >= cached_total:
        required = cached_total
    else:
        required = None

    head_check = needs_head_check(
        args,
        len(cached_videos),
        required,
        cached_total,
        cache.get("continuation_token"),
    )

    if required is not None and len(cached_videos) >= required and not head_check:
        print(
            f"Using cached browse data ({cache_status_line(len(cached_videos), cache, required)}). "
            f"No API pages fetched.",
            flush=True,
        )
        return FetchResult(cached_videos, channel_name, set(), True)

    first_page: list[dict] = []
    first_body: dict | None = None
    new_ids: set[str] = set()
    resume_token: str | None = cache.get("continuation_token")
    resume_page = int(cache.get("pages_fetched") or 0)

    if head_check:
        print("Smart refresh: checking channel head (page 1)...", flush=True)
        browse_api_header()
        cached_ids = {v["id"] for v in cached_videos}
        first_page, channel_name, first_body = fetch_first_browse_page(
            channel_id, known_ids=cached_ids, announce=False
        )
        if cached_videos and not cache_head_matches(first_page, cached_videos):
            print("Cache head mismatch; performing full refresh.", flush=True)
            videos, channel_name, _, pages_fetched = fetch_all_browse_videos(channel_id)
            write_cache(
                cache_path,
                channel_id=channel_id,
                channel_name=channel_name,
                channel_handle=handle or cache.get("channel_handle"),
                videos=videos,
                continuation_token=None,
                pages_fetched=pages_fetched,
            )
            return FetchResult(videos, channel_name, set(), False)
        new_head = new_videos_at_head(first_page, cached_videos) if cached_videos else []
        new_ids = {v["id"] for v in new_head}
        if new_head:
            cached_videos = merge_new_with_cache(new_head, cached_videos)
            print(f"Cache head OK; {len(new_head)} new video(s) at channel head.", flush=True)
            resume_token = None
            resume_page = 0
        elif required is not None and len(cached_videos) >= required:
            print(
                f"Using cached browse data ({cache_status_line(len(cached_videos), cache, required)}).",
                flush=True,
            )
            return FetchResult(cached_videos, channel_name, new_ids, True)

    if required is not None and len(cached_videos) >= required:
        if new_ids:
            write_cache(
                cache_path,
                channel_id=channel_id,
                channel_name=channel_name,
                channel_handle=handle or cache.get("channel_handle"),
                videos=cached_videos,
                continuation_token=resume_token,
                pages_fetched=resume_page or 1,
                keep_playlist_map=True,
            )
        print(
            f"Using cached browse data ({cache_status_line(len(cached_videos), cache, required)}).",
            flush=True,
        )
        return FetchResult(cached_videos, channel_name, new_ids, True)

    if required is not None:
        target_pages = pages_for_index(required)
        cached_pages = int(cache.get("pages_fetched") or 0)
        print(
            f"Smart refresh: {cache_status_line(len(cached_videos), cache, required)}, "
            f"target pages 1-{target_pages}...",
            flush=True,
        )
        if not head_check:
            browse_api_header()
        videos, channel_name, next_token, pages_fetched = fetch_browse_up_to_count(
            channel_id,
            required,
            cached_videos,
            first_body if head_check else None,
            resume_token=resume_token if not new_ids else None,
            resume_after_page=0 if new_ids else resume_page,
        )
    else:
        cached_pages = int(cache.get("pages_fetched") or 0)
        print(
            f"Smart refresh: loading full channel list "
            f"({len(cached_videos)} in cache, pages 1-{cached_pages or '?'})...",
            flush=True,
        )
        if not head_check:
            browse_api_header()
        videos, channel_name, next_token, pages_fetched = fetch_browse_up_to_count(
            channel_id,
            None,
            cached_videos,
            first_body if head_check else None,
            resume_token=resume_token if not new_ids else None,
            resume_after_page=0 if new_ids else resume_page,
        )

    write_cache(
        cache_path,
        channel_id=channel_id,
        channel_name=channel_name,
        channel_handle=handle or cache.get("channel_handle"),
        videos=videos,
        continuation_token=next_token,
        pages_fetched=pages_fetched,
        keep_playlist_map=True,
    )
    return FetchResult(videos, channel_name, new_ids, True)


def load_browse_videos(
    channel_id: str,
    args: argparse.Namespace,
    cache: dict | None,
    handle: str | None,
) -> FetchResult:
    cache_path = cache_file_path(channel_id, args.workspace)
    has_cache = bool(cache and cache.get("videos"))

    if not has_cache:
        if args.refresh or args.new:
            print("No cache yet; --refresh/--new ignored.", flush=True)
        if range_explicit(args) and not args.refresh:
            required = args.to_index
            pages = pages_for_index(required)
            print(
                f"No cache yet; loading pages 1-{pages} ({required} videos needed)...",
                flush=True,
            )
            browse_api_header()
            body = fetch_browse_page(channel_id)
            channel_name = extract_channel_name(body)
            videos, channel_name, next_token, pages_fetched = fetch_browse_up_to_count(
                channel_id, required, [], body
            )
        else:
            videos, channel_name, next_token, pages_fetched = fetch_all_browse_videos(
                channel_id
            )
        write_cache(
            cache_path,
            channel_id=channel_id,
            channel_name=channel_name,
            channel_handle=handle or (cache.get("channel_handle") if cache else None),
            videos=videos,
            continuation_token=next_token,
            pages_fetched=pages_fetched,
        )
        return FetchResult(videos, channel_name, set(), False)

    cached_videos: list[dict] = cache["videos"]
    channel_name = cache.get("channel_name") or fetch_channel_name(channel_id)

    if args.refresh:
        videos, channel_name, _, pages_fetched = fetch_all_browse_videos(channel_id)
        write_cache(
            cache_path,
            channel_id=channel_id,
            channel_name=channel_name,
            channel_handle=handle or cache.get("channel_handle"),
            videos=videos,
            continuation_token=None,
            pages_fetched=pages_fetched,
        )
        return FetchResult(videos, channel_name, set(), False)

    return incremental_smart_load(channel_id, args, cache, handle, cache_path)


def range_explicit(args: argparse.Namespace) -> bool:
    return args.from_index != DEFAULT_FROM or args.to_index != DEFAULT_TO


def selected_channel_indices(
    total: int,
    args: argparse.Namespace,
    new_video_ids: set[str],
    browse_videos: list[dict],
) -> list[int]:
    if args.new and not range_explicit(args):
        indices = [
            i
            for i, video in enumerate(browse_videos, start=1)
            if video["id"] in new_video_ids
        ]
        return indices

    indices: set[int] = set()
    if args.new:
        for i, video in enumerate(browse_videos, start=1):
            if video["id"] in new_video_ids:
                indices.add(i)

    start = max(1, args.from_index)
    end = min(args.to_index, total)
    if start <= end:
        indices.update(range(start, end + 1))
    return sorted(indices)


def build_display_title(title: str, channel_name: str) -> str:
    if re.search(r"\|\s*" + re.escape(channel_name) + r"\s*$", title):
        return title
    return f"{title} | {channel_name}"


def split_display_title(display_title: str, channel_name: str) -> tuple[str, str]:
    match = re.search(r"\s+\|\s+" + re.escape(channel_name) + r"\s*$", display_title)
    if match:
        return display_title[: match.start()].strip(), channel_name
    return display_title.strip(), channel_name


def format_txt_line(record: VideoRecord) -> str:
    display_title = build_display_title(record.title, record.channel_name)
    return (
        f"{record.url} ({display_title} ) {record.date_text} {record.duration_bracket}"
    )


def build_records(
    browse_videos: list[dict],
    channel_name: str,
    selected_indices: list[int],
    playlist_map: dict[str, str],
    today: date,
) -> tuple[list[VideoRecord], int, int]:
    total_channel = len(browse_videos)
    records: list[VideoRecord] = []
    excluded = 0

    for index in selected_indices:
        if index < 1 or index > total_channel:
            continue
        entry = browse_videos[index - 1]
        duration_seconds = duration_text_to_seconds(entry.get("duration_text"))
        date_text = resolve_date(entry["title"], entry.get("relative_published"), today)
        if not should_include(entry["title"], date_text, duration_seconds):
            excluded += 1
            continue
        records.append(
            VideoRecord(
                channel_index=index,
                video_id=entry["id"],
                title=entry["title"].strip(),
                channel_name=channel_name,
                url=entry["url"],
                date_text=date_text,
                duration_bracket=parse_duration_text(entry.get("duration_text")),
                duration_seconds=duration_seconds,
                playlist=playlist_map.get(entry["id"], ""),
            )
        )

    return records, total_channel, excluded


def summary_line(indices: list[int], total_channel: int) -> str:
    if not indices:
        return f"Videos none selected from total {total_channel} videos."
    if len(indices) == 1:
        return f"Video {indices[0]} from total {total_channel} videos."
    return (
        f"Videos {indices[0]} to {indices[-1]} "
        f"({len(indices)} selected) from total {total_channel} videos."
    )


NO_NEW_VIDEOS_LINE = "No new videos since the previous export."


def mode_suffix(args: argparse.Namespace) -> str:
    if args.refresh:
        return "_r"
    if args.new:
        return "_n"
    return ""


def apply_mode_suffix(stem: Path, args: argparse.Namespace) -> Path:
    suffix = mode_suffix(args)
    if suffix:
        return stem.parent / f"{stem.name}{suffix}"
    return stem


def output_stem(
    base_path: Path,
    indices: list[int],
    total_channel: int,
    args: argparse.Namespace,
) -> Path:
    if len(indices) == 1:
        stem = base_path.parent / f"{base_path.name}_{indices[0]}"
    elif len(indices) >= 2 and not (
        indices[0] == 1 and indices[-1] >= total_channel and len(indices) == total_channel
    ):
        stem = base_path.parent / f"{base_path.name}_{indices[0]}_{indices[-1]}"
    else:
        stem = base_path
    return apply_mode_suffix(stem, args)


def write_special_export(stem: Path, line: str) -> tuple[Path, Path]:
    txt_path = stem.with_suffix(".txt")
    xlsx_path = stem.with_suffix(".xlsx")
    txt_path.write_text(line + "\n", encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Videos"
    ws.cell(row=1, column=1, value=line)
    wb.save(xlsx_path)
    return txt_path, xlsx_path


def parse_excel_date(value: str) -> date | None:
    if value == "????-??-??":
        return None
    if len(value) == 7:
        return date(int(value[:4]), int(value[5:7]), 1)
    return date.fromisoformat(value)


def write_txt(path: Path, summary: str, records: list[VideoRecord]) -> None:
    lines = [summary, * (format_txt_line(record) for record in records)]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_xlsx(path: Path, summary: str, records: list[VideoRecord]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Videos"
    ws.cell(row=1, column=1, value=summary)

    row = 2
    for record in records:
        parsed_date = parse_excel_date(record.date_text)
        ws.cell(row=row, column=1, value=record.channel_index)
        ws.cell(row=row, column=2, value=record.channel_name)
        ws.cell(row=row, column=3, value=record.playlist)
        ws.cell(row=row, column=4, value=record.url)
        date_cell = ws.cell(row=row, column=5, value=parsed_date)
        if parsed_date is not None:
            date_cell.number_format = (
                "yyyy-mm-dd" if len(record.date_text) == 10 else "yyyy-mm"
            )
        duration_seconds = record.duration_seconds or duration_bracket_to_seconds(
            record.duration_bracket
        )
        if duration_seconds is not None:
            duration_cell = ws.cell(row=row, column=6, value=duration_seconds / 86400)
            duration_cell.number_format = "[h]:mm:ss"
        display_title = build_display_title(record.title, record.channel_name)
        title_only, _ = split_display_title(display_title, record.channel_name)
        ws.cell(row=row, column=7, value=title_only)
        row += 1

    wb.save(path)


def resolve_playlist_map(
    channel_id: str,
    args: argparse.Namespace,
    cache: dict | None,
    force_refresh: bool,
) -> dict[str, str]:
    if not force_refresh and cache and cache.get("playlist_map"):
        pmap = cache["playlist_map"]
        print(f"Using cached playlist map ({len(pmap)} entries).", flush=True)
        return pmap
    print("Fetching playlists via yt-dlp...", flush=True)
    try:
        return build_video_playlist_map(
            channel_id, lambda *a, **k: yt_dlp_run(args, *a, **k)
        )
    except RuntimeError as exc:
        raise SystemExit(
            f"Playlist lookup failed (playlist column is required):\n{exc}"
        ) from exc


def patch_cache_playlist_map(path: Path, playlist_map: dict[str, str]) -> None:
    cache = read_cache(path)
    if not cache:
        return
    cache["playlist_map"] = playlist_map
    cache["playlist_map_at"] = datetime.now().isoformat(timespec="seconds")
    path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export YouTube channel video summaries to TXT and XLSX."
    )
    parser.add_argument(
        "channel",
        help="YouTube channel id (UC…), @handle, or channel URL",
    )
    parser.add_argument(
        "output",
        nargs="?",
        default=None,
        help="Optional output folder name under output/ (default: derived from @handle)",
    )
    parser.add_argument(
        "--from",
        dest="from_index",
        type=int,
        default=DEFAULT_FROM,
        help=f"First video number on channel (default: {DEFAULT_FROM})",
    )
    parser.add_argument(
        "--to",
        dest="to_index",
        type=int,
        default=DEFAULT_TO,
        help=f"Last video number inclusive (default: {DEFAULT_TO})",
    )
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="Ignore cache and reload all channel data from YouTube",
    )
    parser.add_argument(
        "--new",
        action="store_true",
        help="Include only videos published since the previous export "
        "(unless --from/--to are also set)",
    )
    parser.add_argument(
        "--workspace",
        type=Path,
        default=WORKSPACE_ROOT,
        help="Workspace root (default: parent of src/)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Parent directory for channel folders (default: <workspace>/output)",
    )
    parser.add_argument(
        "--cookies-from-browser",
        metavar="BROWSER",
        help="Pass through to yt-dlp for playlist lookup (chrome, edge, firefox)",
    )
    parser.add_argument(
        "--yt-dlp",
        default="yt-dlp",
        help="yt-dlp executable (default: yt-dlp)",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    today = date.today()

    channel_id, handle = resolve_channel_id(args.channel, args)
    cache_path = cache_file_path(channel_id, args.workspace)
    cache = read_cache(cache_path)

    fetch = load_browse_videos(channel_id, args, cache, handle)
    browse_videos = fetch.videos
    channel_name = fetch.channel_name
    cache = read_cache(cache_path)

    force_playlists = args.refresh or not (cache and cache.get("playlist_map"))
    playlist_map = resolve_playlist_map(channel_id, args, cache, force_playlists)

    folder_name = resolve_output_folder(
        channel_id, handle, args.output, channel_name, args
    )
    out_base = (args.output_dir or output_dir(args.workspace)) / folder_name
    out_base.mkdir(parents=True, exist_ok=True)

    selected = selected_channel_indices(
        len(browse_videos), args, fetch.new_video_ids, browse_videos
    )
    if not selected:
        if args.new and not range_explicit(args):
            stem = apply_mode_suffix(out_base / folder_name, args)
            txt_path, xlsx_path = write_special_export(stem, NO_NEW_VIDEOS_LINE)
            print(f"Channel: {channel_name} ({channel_id})", flush=True)
            if handle:
                print(f"Handle: {handle}", flush=True)
            print(f"Output folder: {out_base}", flush=True)
            print(f"TXT:   {txt_path}", flush=True)
            print(f"XLSX:  {xlsx_path}", flush=True)
            return 0
        print("No videos selected for export.", flush=True)
        return 0

    records, total_channel, excluded = build_records(
        browse_videos, channel_name, selected, playlist_map, today
    )
    summary = summary_line(selected, total_channel)
    stem = output_stem(out_base / folder_name, selected, total_channel, args)
    txt_path = stem.with_suffix(".txt")
    xlsx_path = stem.with_suffix(".xlsx")

    write_txt(txt_path, summary, records)
    write_xlsx(xlsx_path, summary, records)
    if force_playlists:
        patch_cache_playlist_map(cache_path, playlist_map)

    print(f"Channel: {channel_name} ({channel_id})", flush=True)
    if handle:
        print(f"Handle: {handle}", flush=True)
    print(f"Output folder: {out_base}", flush=True)
    print(f"Selected indices: {len(selected)}", flush=True)
    print(f"Exported videos: {len(records)}", flush=True)
    if excluded:
        print(
            f"Excluded from export: {excluded} "
            f"(pre-2020 streams filtered; {total_channel} in browse cache).",
            flush=True,
        )
    print(f"TXT:   {txt_path}", flush=True)
    print(f"XLSX:  {xlsx_path}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
