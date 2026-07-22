#!/usr/bin/env python3
"""Extract slide images from slide-based videos (PySceneDetect).

Videos live in _channels/<channel>/_playlists/<playlist>/ and are processed
in file-name order, like transcribe_videos.py. For each video a scene change
(slide flip) is detected and one image per scene is saved to
<playlist>/SLIDES/<short-key>/slide_NNN.png together with scenes.csv and
scenes.xlsx (scene numbers and timecodes, useful to align slides with .srt
transcripts).

Detection details (tuned on real lecture recordings):

* Change detection is restricted to the central part of the frame
  (--crop-margin), so static screen-share chrome (browser UI, participant
  panel, clock) does not dilute real slide changes. This allows a lower
  threshold that catches subtle slide flips without noise.
* The slide image is captured right after the scene start (--start-offset),
  so every slide_NNN.png matches the Start Timecode of scene NNN in
  scenes.csv (previously the middle frame of the scene was saved, which
  could show a different slide).
* Consecutive scenes whose captured frames are nearly identical in the
  central region (window switches, scrolling, popups that revert) are merged
  into one scene with one image. A scene that shows the same slide merely
  shifted or rescaled (scrolled canvas, editor canvas expanded to full
  screen) is also merged, detected by comparing the frames after aligning
  them with ORB feature matching. Disable with --no-merge.

The short key is the shortest leading substring of the video file stem that
uniquely identifies that video among all videos in the playlist (often just
the leading index, e.g. "01", "02"). This keeps paths short enough for
Windows tools to open scenes.csv reliably.

The session start point is found by scanning SLIDES/: the first video without
a result folder is processed first. Images are written to a temporary folder
and renamed on completion, so an interrupted video is redone on the next run.

Examples:
  python src/extract_slides.py _Autotesting lectures
  python src/extract_slides.py _Autotesting lectures --next 3 --threshold 20
"""

from __future__ import annotations

import argparse
import math
import re
import shutil
import sys
import time
from pathlib import Path

import cv2
import numpy as np

_SRC_DIR = Path(__file__).resolve().parent
if str(_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(_SRC_DIR))

from project_paths import WORKSPACE_ROOT, channels_dir
from transcribe_videos import list_videos

from scenedetect import ContentDetector, SceneManager, open_video
from scenedetect.scene_manager import write_scene_list

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

SLIDES_DIRNAME = "SLIDES"
# Threshold is applied to the cropped central region (see DEFAULT_CROP_MARGIN):
# on lecture screen shares real slide flips score 30+ there while noise stays
# around 10, so 20 catches subtle changes missed by the old full-frame 27.
DEFAULT_THRESHOLD = 20.0
DEFAULT_MIN_SCENE_SECONDS = 4.0
DEFAULT_CROP_MARGIN = 0.15
DEFAULT_START_OFFSET = 1.0
PROGRESS_INTERVAL = 10.0
INVALID_FOLDER_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

# Near-duplicate merging, tuned on lecture recordings. Two-step comparison
# of the central regions of the captured frames:
# 1) pixel-identical check (fast path for a literally unchanged screen);
# 2) edge-map check: Canny edges compared with a small spatial tolerance,
#    so tiny scrolls/jitter do not count, while a new line of text on the
#    slide produces new edges and keeps the scene. Measured values: frames
#    of the same slide differ by 0-8% of edges, different slides by 74-91%.
DEDUP_COMPARE_WIDTH = 192
DEDUP_PIXEL_DELTA = 20
DEDUP_CHANGED_FRACTION = 0.005
EDGE_COMPARE_WIDTH = 960
EDGE_CANNY_LOW, EDGE_CANNY_HIGH = 80, 160
EDGE_TOLERANCE_KERNEL = 7  # dilation kernel: edges may move a few pixels
EDGE_MIN_COUNT = 300       # too few edges -> edge stats meaningless
EDGE_NEW_FRACTION = 0.12
# Same slide rescaled (editor canvas expanded to full screen): ORB feature
# matches estimate the scale transform, frames are compared after alignment.
# Warping blurs edges, so this path uses a looser edge threshold (measured:
# same rescaled slide ~35%, different slides ~85% when a transform exists
# at all - for genuinely different slides RANSAC usually finds none).
ORB_FEATURES = 1500
ORB_MIN_INLIERS = 25
ORB_SCALE_RANGE = (0.2, 5.0)
WARPED_EDGE_NEW_FRACTION = 0.5
WARP_MIN_COVERAGE = 0.2
# Scenes are compared by their "settled" frame taken shortly before the
# scene end: by then slide-flip animations, page loads and transient
# browser toasts ("To exit full screen, press Esc") are over, so the frame
# shows what the scene actually settled on.
SETTLED_TAIL_SECONDS = 2.0


def sanitize_folder_name(name: str) -> str:
    """Make a string safe as a Windows folder name."""
    cleaned = INVALID_FOLDER_CHARS.sub("_", name).rstrip(" .")
    return cleaned or "_"


def short_slide_keys(stems: list[str]) -> dict[str, str]:
    """Map each video stem to the shortest unique leading substring.

    Among all stems in a playlist, find the minimal prefix length such that
    the sanitized prefixes remain unique. Typical result for numbered files
    like ``01_Introduction.mp4`` / ``02_Locators.mp4`` is ``{"01_…": "01", …}``.
    """
    if not stems:
        return {}
    if len(set(stems)) != len(stems):
        raise ValueError("Video stems in the playlist are not unique")

    max_len = max(len(stem) for stem in stems)
    for length in range(1, max_len + 1):
        keys = [sanitize_folder_name(stem[:length]) for stem in stems]
        if len(set(keys)) == len(stems) and all(keys):
            return dict(zip(stems, keys))
    return {stem: sanitize_folder_name(stem) for stem in stems}


def slides_out_dir(slides_dir: Path, video: Path, keys: dict[str, str]) -> Path:
    return slides_dir / keys[video.stem]


def is_extracted(video: Path, slides_dir: Path, keys: dict[str, str]) -> bool:
    out_dir = slides_out_dir(slides_dir, video, keys)
    return out_dir.is_dir() and any(out_dir.iterdir())


def rename_existing_slide_folders(
    playlist_dir: Path, videos: list[Path]
) -> list[tuple[str, str]]:
    """Rename legacy SLIDES/<full-stem>/ folders to SLIDES/<short-key>/.

    Also accepts already-short folders that match the new key (no-op).
    Returns list of (old_name, new_name) renames performed.
    """
    slides_dir = playlist_dir / SLIDES_DIRNAME
    if not slides_dir.is_dir():
        return []
    keys = short_slide_keys([video.stem for video in videos])
    renames: list[tuple[str, str]] = []
    for video in videos:
        short = keys[video.stem]
        target = slides_dir / short
        candidates = [
            slides_dir / video.stem,
            slides_dir / sanitize_folder_name(video.stem),
        ]
        for source in candidates:
            if not source.is_dir() or source.resolve() == target.resolve():
                continue
            if target.exists():
                raise SystemExit(
                    f"Cannot rename {source.name!r} -> {short!r}: "
                    f"target already exists in {slides_dir}"
                )
            source.rename(target)
            renames.append((source.name, short))
            break
    return renames


class ScanProgressStream:
    """Wraps a VideoStream and prints scanning progress every 10 seconds.

    Unlike the built-in tqdm bar, this prints whole lines, so it is equally
    readable in the console and in bat logs. The percentage is real: it is
    the fraction of video frames already scanned.
    """

    def __init__(self, stream, interval: float = PROGRESS_INTERVAL):
        self._stream = stream
        self._interval = interval
        self._frames_read = 0
        self._total_frames = max(1, stream.duration.frame_num)
        self._last_report = time.monotonic()

    def read(self, *args, **kwargs):
        frame = self._stream.read(*args, **kwargs)
        self._frames_read += 1
        now = time.monotonic()
        if now - self._last_report >= self._interval:
            self._last_report = now
            percent = min(99, 100 * self._frames_read // self._total_frames)
            position = self._stream.position.get_timecode()[:8]
            print(f"    Scanning: {percent}%  ({position})", flush=True)
        return frame

    def __getattr__(self, name):
        return getattr(self._stream, name)


def scene_fps(scene: tuple) -> float | None:
    """Frame rate computed from one scene: length in frames / length in seconds."""
    start, end = scene
    frames = end.frame_num - start.frame_num
    seconds = end.seconds - start.seconds
    if frames <= 0 or seconds <= 0:
        return None
    return frames / seconds


def append_fps_column(csv_path: Path, scenes: list[tuple]) -> None:
    """Add a computed 'Frame Rate (fps)' column to scenes.csv, separated from
    the last standard column by one empty column."""
    lines = csv_path.read_text(encoding="utf-8").splitlines()
    out: list[str] = []
    scene_index = 0
    in_scene_rows = False
    for line in lines:
        if line.startswith("Scene Number"):
            out.append(line + ",,Frame Rate (fps)")
            in_scene_rows = True
        elif in_scene_rows and line.strip() and scene_index < len(scenes):
            fps = scene_fps(scenes[scene_index])
            scene_index += 1
            out.append(line + (f",,{fps:.4f}" if fps is not None else ",,"))
        else:
            out.append(line)
    csv_path.write_text("\n".join(out) + "\n", encoding="utf-8")


def central_crop_rect(
    frame_size: tuple[int, int], margin: float
) -> tuple[int, int, int, int] | None:
    """Inclusive (x0, y0, x1, y1) rectangle of the central frame region."""
    if margin <= 0:
        return None
    width, height = frame_size
    x0 = int(width * margin)
    y0 = int(height * margin)
    return (x0, y0, width - 1 - x0, height - 1 - y0)


def capture_frame_at(stream, seconds: float):
    """Read one frame at the given position; None if it cannot be decoded."""
    try:
        stream.seek(max(0.0, seconds))
        frame = stream.read()
    except Exception:
        return None
    return frame if isinstance(frame, np.ndarray) else None


def capture_scene_frame(stream, scene: tuple, offset_seconds: float):
    """Frame shortly after the scene start (matches Start Timecode).

    The offset skips flip animations; for very short scenes it is clamped to
    the scene middle.
    """
    start, end = scene
    length = max(0.0, end.seconds - start.seconds)
    return capture_frame_at(
        stream, start.seconds + min(offset_seconds, length / 2.0)
    )


def _crop_gray(frame, rect: tuple[int, int, int, int] | None):
    if rect is not None:
        x0, y0, x1, y1 = rect
        frame = frame[y0 : y1 + 1, x0 : x1 + 1]
    return cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)


def _changed_fraction(gray_a, gray_b, valid_mask=None) -> float:
    """Fraction of pixels that differ noticeably, at reduced resolution."""
    scale = DEDUP_COMPARE_WIDTH / gray_a.shape[1]
    size = (DEDUP_COMPARE_WIDTH, max(1, round(gray_a.shape[0] * scale)))
    a = cv2.resize(gray_a, size, interpolation=cv2.INTER_AREA)
    b = cv2.resize(gray_b, size, interpolation=cv2.INTER_AREA)
    diff = cv2.absdiff(a, b)
    if valid_mask is None:
        return float(np.count_nonzero(diff > DEDUP_PIXEL_DELTA)) / diff.size
    valid = cv2.resize(valid_mask, size, interpolation=cv2.INTER_NEAREST) > 0
    total = int(np.count_nonzero(valid))
    if total == 0:
        return 1.0
    changed = int(np.count_nonzero((diff > DEDUP_PIXEL_DELTA) & valid))
    return changed / total


def _edge_new_fraction(gray_a, gray_b, valid_mask=None) -> float:
    """Largest one-sided fraction of edges absent from the other frame.

    Edge maps are compared with a small spatial tolerance (dilation), so
    sub-pixel jitter and slight scrolls are ignored while genuinely new
    content (a new line of text, another slide) scores high.
    """
    scale = EDGE_COMPARE_WIDTH / gray_a.shape[1]
    size = (EDGE_COMPARE_WIDTH, max(1, round(gray_a.shape[0] * scale)))
    a = cv2.resize(gray_a, size, interpolation=cv2.INTER_AREA)
    b = cv2.resize(gray_b, size, interpolation=cv2.INTER_AREA)
    edges_a = cv2.Canny(a, EDGE_CANNY_LOW, EDGE_CANNY_HIGH)
    edges_b = cv2.Canny(b, EDGE_CANNY_LOW, EDGE_CANNY_HIGH)
    if valid_mask is not None:
        valid = cv2.resize(valid_mask, size, interpolation=cv2.INTER_NEAREST)
        # Shrink the mask so the border of the warped area itself does not
        # register as an edge.
        valid = cv2.erode(
            valid, np.ones((EDGE_TOLERANCE_KERNEL,) * 2, np.uint8)
        )
        edges_a &= valid
        edges_b &= valid
    count_a = int(np.count_nonzero(edges_a))
    count_b = int(np.count_nonzero(edges_b))
    if min(count_a, count_b) < EDGE_MIN_COUNT:
        # Not enough structure to compare reliably; report "different" so
        # the scene is kept rather than silently merged.
        return 1.0
    kernel = np.ones((EDGE_TOLERANCE_KERNEL,) * 2, np.uint8)
    fat_a = cv2.dilate(edges_a, kernel)
    fat_b = cv2.dilate(edges_b, kernel)
    new_in_b = np.count_nonzero(edges_b & ~fat_a) / count_b
    new_in_a = np.count_nonzero(edges_a & ~fat_b) / count_a
    return max(new_in_a, new_in_b)


def same_screen(frame_a, frame_b, rect) -> bool:
    """True when the central regions show the same slide (small jitter,
    cursor moves and scrolling within the slide are tolerated)."""
    gray_a = _crop_gray(frame_a, rect)
    gray_b = _crop_gray(frame_b, rect)
    if gray_a.shape != gray_b.shape:
        return False
    if _changed_fraction(gray_a, gray_b) < DEDUP_CHANGED_FRACTION:
        return True
    return _edge_new_fraction(gray_a, gray_b) < EDGE_NEW_FRACTION


def same_slide_transformed(frame_a, frame_b, rect) -> bool:
    """True when both frames show the same slide at a different scale.

    Typical case: the slide editor canvas expanded to full screen (or back).
    ORB keypoints of the central regions are matched to estimate the
    geometric transform; edge maps are then compared after alignment, so
    frames with genuinely different content stay separate.
    """
    gray_a = _crop_gray(frame_a, rect)
    gray_b = _crop_gray(frame_b, rect)
    orb = cv2.ORB_create(nfeatures=ORB_FEATURES)
    kp_a, des_a = orb.detectAndCompute(gray_a, None)
    kp_b, des_b = orb.detectAndCompute(gray_b, None)
    if des_a is None or des_b is None:
        return False
    matcher = cv2.BFMatcher(cv2.NORM_HAMMING)
    pairs = matcher.knnMatch(des_a, des_b, k=2)
    good = [m for m, n in (p for p in pairs if len(p) == 2)
            if m.distance < 0.75 * n.distance]
    if len(good) < ORB_MIN_INLIERS:
        return False
    src = np.float32([kp_a[m.queryIdx].pt for m in good]).reshape(-1, 1, 2)
    dst = np.float32([kp_b[m.trainIdx].pt for m in good]).reshape(-1, 1, 2)
    matrix, mask = cv2.findHomography(src, dst, cv2.RANSAC, 5.0)
    if matrix is None or mask is None or int(mask.sum()) < ORB_MIN_INLIERS:
        return False
    det = abs(float(np.linalg.det(matrix[:2, :2])))
    if det <= 0:
        return False
    scale = math.sqrt(det)
    if not (ORB_SCALE_RANGE[0] < scale < ORB_SCALE_RANGE[1]):
        return False
    height, width = gray_b.shape
    warped = cv2.warpPerspective(gray_a, matrix, (width, height))
    coverage = cv2.warpPerspective(
        np.full_like(gray_a, 255), matrix, (width, height)
    )
    if np.count_nonzero(coverage) < WARP_MIN_COVERAGE * coverage.size:
        return False
    return (
        _edge_new_fraction(warped, gray_b, valid_mask=coverage)
        < WARPED_EDGE_NEW_FRACTION
    )


def capture_settled_frame(stream, scene: tuple):
    """Frame SETTLED_TAIL_SECONDS before the scene end (what it settled on)."""
    start, end = scene
    return capture_frame_at(
        stream, max(start.seconds, end.seconds - SETTLED_TAIL_SECONDS)
    )


def merge_similar_scenes(
    stream,
    scenes: list[tuple],
    rect: tuple[int, int, int, int] | None,
    offset_seconds: float,
    *,
    merge: bool,
) -> tuple[list[tuple], list]:
    """Merge consecutive scenes that keep showing the same content.

    Each kept scene is represented by its saved image (start-anchored frame
    of its first sub-scene, matching the Start Timecode in scenes.csv). The
    next raw scene is folded into the current kept scene when its settled
    frame (near the scene end) still shows the same slide as that image -
    directly or rescaled. Returns (scenes, images) of equal length.
    """
    kept_scenes: list[tuple] = []
    kept_images: list = []
    for scene in scenes:
        settled = capture_settled_frame(stream, scene)
        if (
            merge
            and kept_scenes
            and settled is not None
            and kept_images[-1] is not None
            and (
                same_screen(kept_images[-1], settled, rect)
                or same_slide_transformed(kept_images[-1], settled, rect)
            )
        ):
            kept_scenes[-1] = (kept_scenes[-1][0], scene[1])
        else:
            kept_scenes.append(scene)
            kept_images.append(
                capture_scene_frame(stream, scene, offset_seconds)
            )
    return kept_scenes, kept_images


def write_scenes_xlsx(xlsx_path: Path, scenes: list[tuple]) -> None:
    """scenes.xlsx twin of scenes.csv with duration-typed timecode cells.

    CSV cannot carry cell formats, so Excel shows "01:01:47.312" as a time of
    day with the hours hidden. Here timecodes are stored as real durations
    (Excel day fractions) formatted [h]:mm:ss.000.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    duration_format = "[h]:mm:ss.000"
    headers = [
        "Scene Number", "Start Frame", "Start Timecode", "Start Time (seconds)",
        "End Frame", "End Timecode", "End Time (seconds)", "Length (frames)",
        "Length (timecode)", "Length (seconds)", "", "Frame Rate (fps)",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scenes"
    sheet.append(headers)
    for number, (start, end) in enumerate(scenes, start=1):
        length_frames = end.frame_num - start.frame_num
        length_seconds = end.seconds - start.seconds
        fps = scene_fps((start, end))
        row = [
            number, start.frame_num + 1, start.seconds / 86400.0,
            round(start.seconds, 3), end.frame_num, end.seconds / 86400.0,
            round(end.seconds, 3), length_frames, length_seconds / 86400.0,
            round(length_seconds, 3), None,
            round(fps, 4) if fps is not None else None,
        ]
        sheet.append(row)
        for column in (3, 6, 9):
            sheet.cell(row=number + 1, column=column).number_format = (
                duration_format
            )
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = (
            max(12, len(header) + 2)
        )
    workbook.save(xlsx_path)


def extract_video_slides(
    video: Path,
    out_dir: Path,
    *,
    threshold: float,
    min_scene_seconds: float,
    image_format: str,
    crop_margin: float,
    start_offset: float,
    merge_duplicates: bool,
) -> int:
    """Detect slide changes and save one image per scene; return scene count."""
    stream = open_video(str(video))
    rect = central_crop_rect(stream.frame_size, crop_margin)
    manager = SceneManager()
    if rect is not None:
        manager.crop = rect
    manager.add_detector(
        ContentDetector(
            threshold=threshold, min_scene_len=f"{min_scene_seconds}s"
        )
    )
    total_minutes = stream.duration.seconds / 60.0
    print(
        f"  Scanning {stream.duration.frame_num} frames"
        f" ({total_minutes:.0f} min of video)...",
        flush=True,
    )
    manager.detect_scenes(ScanProgressStream(stream), show_progress=False)
    raw_scenes = manager.get_scene_list()
    if not raw_scenes:
        # No cuts detected: treat the whole video as a single slide.
        raw_scenes = [(stream.base_timecode, stream.duration)]

    scenes, frames = merge_similar_scenes(
        stream, raw_scenes, rect, start_offset, merge=merge_duplicates
    )
    if len(scenes) != len(raw_scenes):
        print(
            f"  Scenes: {len(raw_scenes)} detected, "
            f"{len(scenes)} after merging near-duplicates.",
            flush=True,
        )

    tmp_dir = out_dir.with_name(out_dir.name + ".tmp")
    shutil.rmtree(tmp_dir, ignore_errors=True)
    tmp_dir.mkdir(parents=True)
    encode_params = (
        [cv2.IMWRITE_PNG_COMPRESSION, 6]
        if image_format == "png"
        else [cv2.IMWRITE_JPEG_QUALITY, 95]
        if image_format == "jpg"
        else [cv2.IMWRITE_WEBP_QUALITY, 95]
    )
    for number, frame in enumerate(frames, start=1):
        if frame is None:
            print(f"  WARNING: could not decode a frame for scene {number}",
                  flush=True)
            continue
        image_path = tmp_dir / f"slide_{number:03d}.{image_format}"
        cv2.imwrite(str(image_path), frame, encode_params)

    csv_path = tmp_dir / "scenes.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as csv_file:
        write_scene_list(csv_file, scenes)
    append_fps_column(csv_path, scenes)
    write_scenes_xlsx(tmp_dir / "scenes.xlsx", scenes)

    shutil.rmtree(out_dir, ignore_errors=True)
    tmp_dir.rename(out_dir)
    return len(scenes)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract slide images from slide-based videos."
    )
    parser.add_argument(
        "channel_folder",
        help="Channel folder name under _channels (e.g. _Autotesting)",
    )
    parser.add_argument(
        "playlist_folder",
        help="Playlist folder name under <channel>/_playlists (e.g. lectures)",
    )
    parser.add_argument(
        "--next",
        dest="next_count",
        type=int,
        default=1,
        metavar="N",
        help="How many videos to process this session (default: 1)",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=DEFAULT_THRESHOLD,
        help=(
            "Scene change sensitivity, lower = more sensitive "
            f"(default: {DEFAULT_THRESHOLD})"
        ),
    )
    parser.add_argument(
        "--min-scene-seconds",
        type=float,
        default=DEFAULT_MIN_SCENE_SECONDS,
        help=(
            "Minimum slide duration in seconds; shorter changes are merged "
            f"(default: {DEFAULT_MIN_SCENE_SECONDS})"
        ),
    )
    parser.add_argument(
        "--crop-margin",
        type=float,
        default=DEFAULT_CROP_MARGIN,
        help=(
            "Fraction of the frame cut off each edge before change "
            "detection, so static UI around the slides is ignored; "
            f"0 disables cropping (default: {DEFAULT_CROP_MARGIN})"
        ),
    )
    parser.add_argument(
        "--start-offset",
        type=float,
        default=DEFAULT_START_OFFSET,
        help=(
            "Capture the slide image this many seconds after the scene "
            f"start (default: {DEFAULT_START_OFFSET})"
        ),
    )
    parser.add_argument(
        "--no-merge",
        action="store_true",
        help="Keep every detected scene; skip near-duplicate merging",
    )
    parser.add_argument(
        "--image-format",
        choices=("png", "jpg", "webp"),
        default="png",
        help="Slide image format (default: png)",
    )
    parser.add_argument(
        "--rename-existing",
        action="store_true",
        help=(
            "Only rename existing SLIDES/<full-stem>/ folders to short unique "
            "keys; do not extract new slides"
        ),
    )
    parser.add_argument(
        "--workspace",
        type=Path,
        default=WORKSPACE_ROOT,
        help="Workspace root (default: parent of src/)",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.next_count < 1:
        raise SystemExit("--next must be a positive number")

    playlist_dir = (
        channels_dir(args.workspace)
        / args.channel_folder
        / "_playlists"
        / args.playlist_folder
    )
    if not playlist_dir.is_dir():
        raise SystemExit(f"Playlist folder not found: {playlist_dir}")

    slides_dir = playlist_dir / SLIDES_DIRNAME
    videos = list_videos(playlist_dir)
    if not videos:
        raise SystemExit(f"No video/audio files found in {playlist_dir}")

    keys = short_slide_keys([video.stem for video in videos])
    sample = next(iter(keys.values()))
    print(f"Playlist folder: {playlist_dir}", flush=True)
    print(
        f"Slide folder keys: length {len(sample)} "
        f"(e.g. {', '.join(list(keys.values())[:3])}…)",
        flush=True,
    )

    if args.rename_existing:
        renames = rename_existing_slide_folders(playlist_dir, videos)
        if not renames:
            print("No SLIDES folders needed renaming.", flush=True)
        else:
            for old, new in renames:
                print(f"  Renamed: {old}  ->  {new}", flush=True)
            print(f"Renamed {len(renames)} folder(s).", flush=True)
        return 0

    pending = [
        video for video in videos if not is_extracted(video, slides_dir, keys)
    ]
    session = pending[: args.next_count]

    print(
        f"Videos: {len(videos)} total, {len(videos) - len(pending)} already done, "
        f"{len(session)} in this session (--next {args.next_count}).",
        flush=True,
    )
    if not session:
        print("Nothing to do: slides are extracted for all videos.", flush=True)
        return 0

    for index, video in enumerate(session, start=1):
        print(f"[{index}/{len(session)}] {video.name}", flush=True)
        out_dir = slides_out_dir(slides_dir, video, keys)
        count = extract_video_slides(
            video,
            out_dir,
            threshold=args.threshold,
            min_scene_seconds=args.min_scene_seconds,
            image_format=args.image_format,
            crop_margin=args.crop_margin,
            start_offset=args.start_offset,
            merge_duplicates=not args.no_merge,
        )
        print(
            f"  Saved: {out_dir.relative_to(playlist_dir)} "
            f"({count} slide(s) + scenes.csv + scenes.xlsx)",
            flush=True,
        )

    print(f"Session done: {len(session)} video(s).", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
