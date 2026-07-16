#!/usr/bin/env python3
"""Generate download_videos.bat from Vladilen_Minin.xlsx (column C URLs)."""

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Vladilen_Minin.xlsx"
BAT = ROOT / "download_videos.bat"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    urls: list[str] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        url = row[2]
        if url and str(url).startswith("http"):
            urls.append(str(url).strip())

    lines = ["@echo off", "cd /d \"%~dp0\"", ""] + [f"yt-dlp {url}" for url in urls]
    BAT.write_text("\r\n".join(lines) + "\r\n", encoding="utf-8")
    print(f"Wrote {len(urls)} commands to {BAT.name}")


if __name__ == "__main__":
    main()
